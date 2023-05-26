import japanize_matplotlib
import matplotlib.pyplot as plt
import matplotlib
import numpy as np
import pandas as pd
from PIL import Image
import openpyxl as px
import streamlit as st
#import os

#------------------家計簿のグラフ＆表表示＆データダウンロード------------------
#os.chdir(r'C:\Users\e12482\Documents\my_python\streamlit4')
#------------------2023年度はじまり------------------
st.set_page_config(layout="wide")
st.title(':blue[家計簿グラフ]:sunglasses:')
wb=px.load_workbook("./sasaki(FY2023).xlsx")


data1=np.zeros((0,10))#(縦×横)10列の白紙のデータを生成
c=0
for k in range(1,13):
    ws=wb.worksheets[k]
    
    list1=[]
    list2=[]
    list3=[]
    

    for s in range(3,12,4):#すべての支払いデータをlistに格納

        t=21
        while ws.cell(row=t,column=s).value!=None:
            t=t+1

        for i in range(22,t):
            value1=ws.cell(row=i,column=s).value
            value2=ws.cell(row=i,column=s+1).value
            list1.append(value1)
            list2.append(value2)


    data=np.array([list1,list2])#すべての支払い額と支払い種類の表をnumpy配列に変換
    df=pd.DataFrame(data)#pandasフレームに変換
    df=df.T#行と列の入れ替え
    df.columns=["値段","種類"]
    df["値段"]=df["値段"].astype(int)


    yatin=ws.cell(row=2,column=2).value
    dennkigasu=ws.cell(row=3,column=2).value
    suidou=ws.cell(row=4,column=2).value
    tyuusyazyou=ws.cell(row=5,column=2).value
    wifi=ws.cell(row=6,column=2).value
    #種類別の合計を求める↓↓
    koutuuhi=df[df["種類"]=="交通費"]["値段"].sum()
    gasorinhi=df[df["種類"]=="ガソリン"]["値段"].sum()
    syokuhi=df[df["種類"]=="食費"]["値段"].sum()
    zakkahi=df[df["種類"]=="雑貨"]["値段"].sum()
    sonotahi=df[df["種類"]=="その他"]["値段"].sum()

    
    data2=np.array([yatin,dennkigasu,suidou,tyuusyazyou,wifi,koutuuhi,gasorinhi,syokuhi,zakkahi,sonotahi])
    data1=np.vstack([data1,data2])#はじめに作った白紙データに格納していく。



df2=pd.DataFrame(data1)#pandasフレームに変換
df2.index=["2023.4月","2023.5月","2023.6月","2023.7月","2023.8月","2023.9月","2023.10月","2023.11月","2023.12月","2024.1月","2024.2月","2024.3月"]
df2.columns=["家賃","電気ガス","水道","駐車場","wifi","交通費","ガソリン","食費","雑貨","その他"] 

df2=df2.fillna(0)#欠損処理(NaNを0にする)
   

#グラフ化↓↓
fig=plt.figure(figsize=(12,8),facecolor='lightblue')#figを定義(グラフのみではなくそれを含む全体)
ax1=fig.add_subplot()#グラフを定義
df3=df2/2 
#積みあげ棒グラフの作成
for i in range(len(df2.columns)):
    ax1.bar(df2.index,df2.iloc[:,i],bottom=df2.iloc[:,:i].sum(axis=1) )
    

    for j in range(0,len(df2.index)):
        
            s=str(df2.columns[i])
            plt.text(j,df2.iloc[j,:i].sum()+df3.iloc[j,i],s,ha='center', va='center')


ax1.set(xlabel='月', ylabel='金額[円]')
ax1.set_title("月別出費推移グラフ(2023年度)",fontsize=20)

plt.show()
st.pyplot(fig)


def color_background_lightgreen(val):
    color = 'lightgreen' if val == "交通費" else '' #1より大なら薄緑、その他は白
    return 'background-color: %s' % color

#表の表示
option=st.selectbox("表示したい月",df2.index)
if option:
     
    df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=option)
    df3["↓↓共用カード払い"]=df3["↓↓共用カード払い"].astype(str)
    df3["↓↓美織払い"]=df3["↓↓美織払い"].astype(str)
    df3["↓↓洋太払い"]=df3["↓↓洋太払い"].astype(str)
    st.dataframe(df3,width=1500)

    
#エクセルファイルのダウンロード
#@st.cache_data
def get_file(path):
    with open(path, 'rb') as f:
        data = f.read()
    return data

path = r'./sasaki(FY2023).xlsx' # Webサービスで作成、保存したファイルパス
data = get_file(path)
st.write("↓↓ここからExcelファイルダウンロード")
st.download_button( label='download(FY2023)', data=data, file_name='dl.xlsx')

#------------------2023年度おわり------------------



#------------------2022年度はじまり------------------

df=pd.read_excel("./sasaki(FY2022).xlsx",index_col=0)


fig=plt.figure(figsize=(12,8),facecolor='lightblue')
ax1=fig.add_subplot()
df2=df/2 
df.index=df.index.astype(str)

for i in range(len(df.columns)):
    ax1.bar(df.index,df.iloc[:,i],bottom=df.iloc[:,:i].sum(axis=1) )
    

    for j in range(0,len(df.index)):
        
            s=str(df.columns[i])
            plt.text(j,df.iloc[j,:i].sum()+df2.iloc[j,i],s,ha='center', va='center')


ax1.set(xlabel='月', ylabel='金額')
ax1.set_title("月別出費推移グラフ(2022年度)",fontsize=20)

plt.show()
st.pyplot(fig)

   
option=st.selectbox("表示したい月",df.index)
if option:
     
    df3=pd.read_excel("./sasaki(FY2022).xlsx",skiprows=20,sheet_name=option)
    df3["↓↓共用カード払い"]=df3["↓↓共用カード払い"].astype(str)
    df3["↓↓美織払い"]=df3["↓↓美織払い"].astype(str)
    df3["↓↓洋太払い"]=df3["↓↓洋太払い"].astype(str)
    st.dataframe(df3,width=1500)

    





path = r'./sasaki(FY2022).xlsx' # Webサービスで作成、保存したファイルパス
data = get_file(path)
st.write("↓↓ここからExcelファイルダウンロード")
st.download_button( label='download(FY2022)', data=data, file_name='dl.xlsx')

#------------------2022年度おわり------------------