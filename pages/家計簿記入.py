import streamlit as st
import time
import datetime
import openpyxl as px
import pandas as pd
#import os
#os.chdir(r'C:\Users\e12482\Documents\my_python\streamlit4')

#------------------エクセルへの記入------------------
st.set_page_config(layout="wide")
st.title(':blue[記入したい項目あったらここからしてね!!]:sunglasses:')

inputDate=st.date_input(label='日付')
inputText = st.text_input(label='内容', value='スーパーアルプス')
inputValue = st.text_input(label='値段', value='100')
inputValue=int(inputValue)
listType=["交通費","ガソリン","食費","雑貨","その他"]
inputType=st.selectbox("分類",listType)
inputHuman1=st.checkbox(label="美織")
inputHuman2=st.checkbox(label="洋太")
submit_btn=st.button("送信")
listDate1=[]
listDate2=[]
wb=px.load_workbook("./sasaki(FY2023).xlsx")
now = datetime.date.today()

#日付期間のリスト作成
for i in range(3,13):
    a=datetime.date(2023,i,16)
    listDate1.append(a)

for i in range(1,3):
    a=datetime.date(2024,i,16)
    listDate1.append(a)


for i in range(4,13):
    a=datetime.date(2023,i,15)
    listDate2.append(a)

for i in range(1,4):
    a=datetime.date(2024,i,15)
    listDate2.append(a)


#美織選択した場合の記入
if submit_btn:
    if inputHuman1:


        if inputDate>listDate1[0] and inputDate<listDate2[0]:
            t=21
            ws=wb.worksheets[1]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[0])
            st.write(listDate2[0])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=1)
            


        elif inputDate>listDate1[1] and inputDate<listDate2[1]:
            t=21
            ws=wb.worksheets[2]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[1])
            st.write(listDate2[1])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=2)
        
            
            
        elif inputDate>listDate1[2] and inputDate<listDate2[2]:
            t=21
            ws=wb.worksheets[3]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[2])
            st.write(listDate2[2])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=3)
            
        
        
        elif inputDate>listDate1[3] and inputDate<listDate2[3]:
            t=21
            ws=wb.worksheets[4]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[3])
            st.write(listDate2[3])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=4)
        
            
        
        elif inputDate>listDate1[4] and inputDate<listDate2[4]:
            t=21
            ws=wb.worksheets[5]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[4])
            st.write(listDate2[4])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=5)
            
            
        
        
        elif inputDate>listDate1[5] and inputDate<listDate2[5]:
            t=21
            ws=wb.worksheets[6]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[5])
            st.write(listDate2[5])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=6)
        
            
            
        elif inputDate>listDate1[6] and inputDate<listDate2[6]:
            t=21
            ws=wb.worksheets[7]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[6])
            st.write(listDate2[6])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=7)
            
            
        
        elif inputDate>listDate1[7] and inputDate<listDate2[7]:
            t=21
            ws=wb.worksheets[8]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[7])
            st.write(listDate2[7])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=8)


        elif inputDate>listDate1[8] and inputDate<listDate2[8]:
            t=21
            ws=wb.worksheets[9]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[8])
            st.write(listDate2[8])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=9)
            
        elif inputDate>listDate1[9] and inputDate<listDate2[9]:
            t=21
            ws=wb.worksheets[10]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[9])
            st.write(listDate2[9])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=10) 


        elif inputDate>listDate1[10] and inputDate<listDate2[10]:
            t=21
            ws=wb.worksheets[11]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[10])
            st.write(listDate2[10])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=11) 

        elif inputDate>listDate1[11] and inputDate<listDate2[11]:
            t=21
            ws=wb.worksheets[12]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[11])
            st.write(listDate2[11])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=12) 
        
        elif inputDate>listDate1[12] and inputDate<listDate2[12]:
            t=21
            ws=wb.worksheets[13]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[12])
            st.write(listDate2[12])
            while ws.cell(row=t,column=5).value!=None:
                t=t+1
            ws.cell(row=t,column=5).value=inputDate
            ws.cell(row=t,column=6).value=inputText
            ws.cell(row=t,column=7).value=inputValue
            ws.cell(row=t,column=8).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=13) 
            
        
        df3["↓↓共用カード払い"]=df3["↓↓共用カード払い"].astype(str)
        df3["↓↓美織払い"]=df3["↓↓美織払い"].astype(str)
        df3["↓↓洋太払い"]=df3["↓↓洋太払い"].astype(str)
        
        st.dataframe(df3,width=1500)

    #洋太選択した場合の記入
    if inputHuman2:


        if inputDate>listDate1[0] and inputDate<listDate2[0]:
            t=21
            ws=wb.worksheets[1]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[0])
            st.write(listDate2[0])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=1)
            


        elif inputDate>listDate1[1] and inputDate<listDate2[1]:
            t=21
            ws=wb.worksheets[2]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[1])
            st.write(listDate2[1])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=2)
        
            
            
        elif inputDate>listDate1[2] and inputDate<listDate2[2]:
            t=21
            ws=wb.worksheets[3]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[2])
            st.write(listDate2[2])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=3)
            
        
        
        elif inputDate>listDate1[3] and inputDate<listDate2[3]:
            t=21
            ws=wb.worksheets[4]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[3])
            st.write(listDate2[3])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=4)
        
            
        
        elif inputDate>listDate1[4] and inputDate<listDate2[4]:
            t=21
            ws=wb.worksheets[5]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[4])
            st.write(listDate2[4])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=5)
            
            
        
        
        elif inputDate>listDate1[5] and inputDate<listDate2[5]:
            t=21
            ws=wb.worksheets[6]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[5])
            st.write(listDate2[5])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=6)
        
            
            
        elif inputDate>listDate1[6] and inputDate<listDate2[6]:
            t=21
            ws=wb.worksheets[7]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[6])
            st.write(listDate2[6])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=7)
            
            
        
        elif inputDate>listDate1[7] and inputDate<listDate2[7]:
            t=21
            ws=wb.worksheets[8]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[7])
            st.write(listDate2[7])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=8)
            
        elif inputDate>listDate1[8] and inputDate<listDate2[8]:
            t=21
            ws=wb.worksheets[9]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[8])
            st.write(listDate2[8])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=9)    
        
        elif inputDate>listDate1[9] and inputDate<listDate2[9]:
            t=21
            ws=wb.worksheets[10]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[9])
            st.write(listDate2[9])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=10)  
            
        elif inputDate>listDate1[10] and inputDate<listDate2[10]:
            t=21
            ws=wb.worksheets[11]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[10])
            st.write(listDate2[10])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=11)  
            

        elif inputDate>listDate1[11] and inputDate<listDate2[11]:
            t=21
            ws=wb.worksheets[12]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[11])
            st.write(listDate2[11])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=12) 

        elif inputDate>listDate1[12] and inputDate<listDate2[12]:
            t=21
            ws=wb.worksheets[13]
            st.write("記入ありがとうございます!!みこちゃま🐵")
            st.write(listDate1[12])
            st.write(listDate2[12])
            while ws.cell(row=t,column=9).value!=None:
                t=t+1
            ws.cell(row=t,column=9).value=inputDate
            ws.cell(row=t,column=10).value=inputText
            ws.cell(row=t,column=11).value=inputValue
            ws.cell(row=t,column=12).value=inputType
            wb.save("./sasaki(FY2023).xlsx")
            df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=13) 
            
        df3["↓↓共用カード払い"]=df3["↓↓共用カード払い"].astype(str)
        df3["↓↓美織払い"]=df3["↓↓美織払い"].astype(str)
        df3["↓↓洋太払い"]=df3["↓↓洋太払い"].astype(str)
        st.dataframe(df3,width=1500)
       











    



