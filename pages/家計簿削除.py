import streamlit as st
import time
import datetime
import openpyxl as px
import pandas as pd
import openpyxl as px
#import os

#os.chdir(r'C:\Users\e12482\Documents\my_python\streamlit5')

#------------------エクセル値の削除------------------
st.set_page_config(layout="wide")
st.title(':blue[削除したい項目あったらここからしてね!!]:sunglasses:')
df=pd.read_excel("./sasaki(FY2023).xlsx",index_col=0)
wb=px.load_workbook("./sasaki(FY2023).xlsx")
mylist=list(range(20))


option1=st.selectbox("表示したい月",df.index)
option2=st.selectbox("削除したい行",mylist)
inputHuman1=st.checkbox(label="美織")
inputHuman2=st.checkbox(label="洋太")
submit_btn=st.button("削除")
#表示したい月の表示
if option1:
     
    df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=option1)
    df3["↓↓共用カード払い"]=df3["↓↓共用カード払い"].astype(str)
    df3["↓↓美織払い"]=df3["↓↓美織払い"].astype(str)
    df3["↓↓洋太払い"]=df3["↓↓洋太払い"].astype(str)
    st.write("削除前↓↓")
    st.dataframe(df3,width=1500)
#削除したいところ
if submit_btn==True:
    for i in range(0,13):
        ws=wb.worksheets[i]

        if ws.title==option1:#シート名と選択した月の一致
            print(ws.title)
            if inputHuman1:
                t=21
                while ws.cell(row=t,column=5).value!=None:
                    t=t+1

                ws.cell(row=option2+22,column=5).value=None
                ws.cell(row=option2+22,column=6).value=None
                ws.cell(row=option2+22,column=7).value=None
                ws.cell(row=option2+22,column=8).value=None
                

                for k in range(option2+23,t):
                    ws.cell(row=k-1,column=5).value=ws.cell(row=k,column=5).value
                    ws.cell(row=k-1,column=6).value=ws.cell(row=k,column=6).value
                    ws.cell(row=k-1,column=7).value=ws.cell(row=k,column=7).value
                    ws.cell(row=k-1,column=8).value=ws.cell(row=k,column=8).value

                ws.cell(row=t-1,column=5).value=None
                ws.cell(row=t-1,column=6).value=None
                ws.cell(row=t-1,column=7).value=None
                ws.cell(row=t-1,column=8).value=None





            elif inputHuman2:
                t=21
                while ws.cell(row=t,column=9).value!=None:
                    t=t+1
                

                ws.cell(row=option2+22,column=9).value=None
                ws.cell(row=option2+22,column=10).value=None
                ws.cell(row=option2+22,column=11).value=None
                ws.cell(row=option2+22,column=12).value=None

                

                for k in range(option2+23,t):
                    ws.cell(row=k-1,column=9).value=ws.cell(row=k,column=9).value
                    ws.cell(row=k-1,column=10).value=ws.cell(row=k,column=10).value
                    ws.cell(row=k-1,column=11).value=ws.cell(row=k,column=11).value
                    ws.cell(row=k-1,column=12).value=ws.cell(row=k,column=12).value

                ws.cell(row=t-1,column=9).value=None
                ws.cell(row=t-1,column=10).value=None
                ws.cell(row=t-1,column=11).value=None
                ws.cell(row=t-1,column=12).value=None
                

    wb.save("./sasaki(FY2023).xlsx")
    df3=pd.read_excel("./sasaki(FY2023).xlsx",skiprows=20,sheet_name=option1)
    df3["↓↓共用カード払い"]=df3["↓↓共用カード払い"].astype(str)
    df3["↓↓美織払い"]=df3["↓↓美織払い"].astype(str)
    df3["↓↓洋太払い"]=df3["↓↓洋太払い"].astype(str)
    st.write("修正ありがとうございます!!みこちゃま🐵")
    st.write("削除後↓↓")
    st.dataframe(df3,width=1500)
  
