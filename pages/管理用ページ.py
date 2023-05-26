import japanize_matplotlib
import matplotlib.pyplot as plt
import matplotlib
import numpy as np
import pandas as pd
from PIL import Image
import openpyxl as px
import streamlit as st
#import os

#------------------ファイルアップロード------------------
#os.chdir(r'C:\Users\e12482\Documents\my_python\streamlit5')
st.set_page_config(layout="wide")
st.title(':blue[ファイルアップロード]:sunglasses:')
df=pd.read_excel("./sasaki(FY2023).xlsx",index_col=0)
wb=px.load_workbook("./sasaki(FY2023).xlsx")
option=st.selectbox("表示したい月",df.index,key=1)

uploaded_file=st.file_uploader("ファイルアップロード", type='csv')

if uploaded_file is not None:
    for i in range(0,13):
        ws=wb.worksheets[i]
        if ws.title==option:#シート名と選択した月の一致
    
            df=pd.read_csv(uploaded_file,encoding="Shift-JIS")
            st.dataframe(df)
            data=df.to_numpy()
            print(data)
            print(data.shape)
            print(data.shape[0])
            print(data[0][0])
            print(data[0][1])
            for k in range(0,data.shape[0]):
                ws.cell(row=k+22,column=1).value=data[k][0]
                ws.cell(row=k+22,column=2).value=data[k][1]
                ws.cell(row=k+22,column=3).value=data[k][2]
                ws.cell(row=k+22,column=4).value=data[k][3]

            wb.save("./sasaki(FY2023).xlsx")
            st.write("完了しました")
        



st.title(':blue[固定費記入]:sunglasses:')

listType1=["2023.4月","2023.5月","2023.6月","2023.7月","2023.8月","2023.9月","2023.10月","2023.11月","2023.12月","2024.1月","2024.2月","2024.3月"]
option1=st.selectbox("表示したい月",listType1,key=2)
listType2=["家賃","電気ガス","水道","駐車場","wifi"]
inputType=st.selectbox("分類",listType2)
inputValue = st.text_input(label='値段', value='100')
submit_btn=st.button("記入")

if submit_btn==True:
    for i in range(0,13):
        ws=wb.worksheets[i]
        if ws.title==option1:#シート名と選択した月の一致
            for c in range(2,7):
                if ws.cell(row=c,column=1).value==inputType:
                    ws.cell(row=c,column=2).value=int(inputValue)

                    wb.save("./sasaki(FY2023).xlsx")
                    st.write("完了しました")
        


